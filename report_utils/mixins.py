from six import BytesIO, text_type, string_types

from django.http import HttpResponse
from django.contrib.contenttypes.models import ContentType
from django.db.models.fields.related import ReverseManyRelatedObjectsDescriptor
from django.db.models import Avg, Count, Sum, Max, Min
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Font
import re
from collections import namedtuple
from decimal import Decimal
from numbers import Number
from functools import reduce
import datetime

from report_utils.model_introspection import (
    get_relation_fields_from_model,
    get_properties_from_model,
    get_direct_fields_from_model,
    get_model_from_path_string,
    get_custom_fields_from_model,
)

DisplayField = namedtuple(
    "DisplayField",
    "path path_verbose field field_verbose aggregate total group choices field_type",
)

class DataExportMixin(object):
    def build_sheet(self, data, ws, sheet_name='report', header=None, widths=None):
        first_row = 1
        column_base = 1

        ws.title = re.sub(r'\W+', '', sheet_name)[:30]
        if header:
            for i, header_cell in enumerate(header):
                cell = ws.cell(row=first_row, column=i+column_base)
                cell.value = header_cell
                cell.font = Font(bold=True)
                if widths:
                    ws.column_dimensions[get_column_letter(i+1)].width = widths[i]

        for row in data:
            for i in range(len(row)):
                item = row[i]
                # If item is a regular string
                if isinstance(item, str):
                    # Change it to a unicode string
                    try:
                        row[i] = text_type(item)
                    except UnicodeDecodeError:
                        row[i] = text_type(item.decode('utf-8', 'ignore'))
                elif type(item) is dict:
                    row[i] = text_type(item)
            try:
                ws.append(row)
            except ValueError as e:
                ws.append([e.message])
            except:
                ws.append(['Unknown Error'])

    def build_xlsx_response(self, wb, title="report"):
        """ Take a workbook and return a xlsx file response """
        if not title.endswith('.xlsx'):
            title += '.xlsx'
        myfile = BytesIO()
        myfile.write(save_virtual_workbook(wb))
        response = HttpResponse(
            myfile.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s' % title
        response['Content-Length'] = myfile.tell()
        return response

    def list_to_workbook(self, data, title='report', header=None, widths=None):
        """ Create just a openpxl workbook from a list of data """
        wb = Workbook()
        title = re.sub(r'\W+', '', title)[:30]

        if isinstance(data, dict):
            i = 0
            for sheet_name, sheet_data in data.items():
                if i > 0:
                    wb.create_sheet()
                ws = wb.worksheets[i]
                self.build_sheet(
                    sheet_data, ws, sheet_name=sheet_name, header=header)
                i += 1
        else:
            ws = wb.worksheets[0]
            self.build_sheet(data, ws, header=header, widths=widths)
        return wb

    def list_to_xlsx_file(self, data, title='report', header=None, widths=None):
        """ Make 2D list into a xlsx response for download
        data can be a 2d array or a dict of 2d arrays
        like {'sheet_1': [['A1', 'B1']]}
        returns a StringIO file
        """
        wb = self.list_to_workbook(data, title, header, widths)
        if not title.endswith('.xlsx'):
            title += '.xlsx'
        myfile = BytesIO()
        myfile.write(save_virtual_workbook(wb))
        return myfile

    def list_to_xlsx_response(self, data, title='report', header=None,
                              widths=None):
        """ Make 2D list into a xlsx response for download
        data can be a 2d array or a dict of 2d arrays
        like {'sheet_1': [['A1', 'B1']]}
        """
        wb = self.list_to_workbook(data, title, header, widths)
        return self.build_xlsx_response(wb, title=title)

    def add_aggregates(self, queryset, display_fields):
        for display_field in display_fields:
            if hasattr(display_field, 'aggregate'):
                if display_field.aggregate == "Avg":
                    queryset = queryset.annotate(Avg(display_field.path + display_field.field))
                elif display_field.aggregate == "Max":
                    queryset = queryset.annotate(Max(display_field.path + display_field.field))
                elif display_field.aggregate == "Min":
                    queryset = queryset.annotate(Min(display_field.path + display_field.field))
                elif display_field.aggregate == "Count":
                    queryset = queryset.annotate(Count(display_field.path + display_field.field))
                elif display_field.aggregate == "Sum":
                    queryset = queryset.annotate(Sum(display_field.path + display_field.field))
        return queryset

    def report_to_list(self, queryset, display_fields, user, property_filters=[], preview=False):
        """ Create list from a report with all data filtering
        preview: Return only first 50
        objects: Provide objects for list, instead of running filters
        display_fields: a list of fields or a report_builder display field model
        Returns list, message in case of issues
        """
        model_class = queryset.model
        if isinstance(display_fields, list):
            # Make it a report_builder.models.DisplayField like object
            new_display_fields = []
            for display_field in display_fields:
                field_list = display_field.split('__')
                field = field_list[-1]
                path = '__'.join([str(x) for x in field_list[:-1]])
                if path:
                    path += '__'  # Legacy format to append a __ here.
                new_model = get_model_from_path_string(model_class, path)
                model_field = new_model._meta.get_field_by_name(field)[0]
                choices = model_field.choices
                new_display_fields.append(DisplayField(
                    path, '', field, '', '', None, None, choices, ''))
            display_fields = new_display_fields

        message = ""
        objects = self.add_aggregates(queryset, display_fields)

        # Display Values
        display_field_paths = []
        property_list = {}
        custom_list = {}
        display_totals = {}

        def append_display_total(display_totals, display_field,
                                 display_field_key):
            if display_field.total:
                display_totals[display_field_key] = {'val': Decimal('0.00')}

        for i, display_field in enumerate(display_fields):
            model = get_model_from_path_string(model_class, display_field.path)
            if display_field.field_type == "Invalid":
                continue
            if (user.has_perm(model._meta.app_label + '.change_' + model._meta.model_name)
                    or user.has_perm(model._meta.app_label + '.view_' + model._meta.model_name)
                    or not model):
                # TODO: clean this up a bit
                display_field_key = display_field.path + display_field.field
                if display_field.field_type == "Property":
                    property_list[i] = display_field_key
                    append_display_total(display_totals, display_field, display_field_key)
                elif display_field.field_type == "Custom Field":
                    custom_list[i] = display_field_key
                    append_display_total(display_totals, display_field, display_field_key)
                elif display_field.aggregate == "Avg":
                    display_field_key += '__avg'
                    display_field_paths += [display_field_key]
                    append_display_total(display_totals, display_field, display_field_key)
                elif display_field.aggregate == "Max":
                    display_field_key += '__max'
                    display_field_paths += [display_field_key]
                    append_display_total(display_totals, display_field, display_field_key)
                elif display_field.aggregate == "Min":
                    display_field_key += '__min'
                    display_field_paths += [display_field_key]
                    append_display_total(display_totals, display_field, display_field_key)
                elif display_field.aggregate == "Count":
                    display_field_key += '__count'
                    display_field_paths += [display_field_key]
                    append_display_total(display_totals, display_field, display_field_key)
                elif display_field.aggregate == "Sum":
                    display_field_key += '__sum'
                    display_field_paths += [display_field_key]
                    append_display_total(display_totals, display_field, display_field_key)
                else:
                    display_field_paths += [display_field_key]
                    append_display_total(display_totals, display_field, display_field_key)
            else:
                message += "You don't have permission to " + display_field.name

        try:
            model_name = model_class._meta.model_name
        except AttributeError:
            model_name = model_class._meta.module_name # needed for Django 1.4.* (LTS)

        if user.has_perm(model_class._meta.app_label + '.change_' + model_name) \
        or user.has_perm(model_class._meta.app_label + '.view_' + model_name):

            def increment_total(display_field_key, display_totals, val):
                if display_field_key in display_totals:
                    if isinstance(val, bool):
                        # True: 1, False: 0
                        display_totals[display_field_key]['val'] += Decimal(val)
                    elif isinstance(val, Number):
                        display_totals[display_field_key]['val'] += Decimal(str(val))
                    elif val:
                        display_totals[display_field_key]['val'] += Decimal(1)

            # get pk for primary and m2m relations in order to retrieve objects
            # for adding properties to report rows
            display_field_paths.insert(0, 'pk')
            m2m_relations = []
            for position, property_path in property_list.items():
                property_root = property_path.split('__')[0]
                root_class = model_class
                try:
                    property_root_class = getattr(root_class, property_root)
                # django-hstore schema compatibility
                except AttributeError:
                    continue
                if type(property_root_class) == ReverseManyRelatedObjectsDescriptor:
                    display_field_paths.insert(1, '%s__pk' % property_root)
                    m2m_relations.append(property_root)
            values_and_properties_list = []
            filtered_report_rows = []
            group = None
            for df in display_fields:
                if df.group:
                    group = df.path + df.field
                    break
            if group:
                filtered_report_rows = self.add_aggregates(objects.values_list(group), display_fields)
            else:
                values_list = objects.values_list(*display_field_paths)

            if not group:
                for row in values_list:
                    row = list(row)
                    values_and_properties_list.append(row[1:])
                    obj = None # we will get this only if needed for more complex processing
                    #related_objects
                    remove_row = False
                    # filter properties (remove rows with excluded properties)
                    for property_filter in property_filters:
                        if not obj:
                            obj = model_class.objects.get(pk=row.pop(0))
                        root_relation = property_filter.path.split('__')[0]
                        if root_relation in m2m_relations:
                            pk = row[0]
                            if pk is not None:
                                # a related object exists
                                m2m_obj = getattr(obj, root_relation).get(pk=pk)
                                val = reduce(getattr, [property_filter.field], m2m_obj)
                            else:
                                val = None
                        else:
                            if property_filter.field_type == 'Custom Field':
                                for relation in property_filter.path.split('__'):
                                    if hasattr(obj, root_relation):
                                        obj = getattr(obj, root_relation)
                                val = obj.get_custom_value(property_filter.field)
                            else:
                                val = reduce(getattr, (property_filter.path + property_filter.field).split('__'), obj)
                        if property_filter.filter_property(val):
                            remove_row = True
                            values_and_properties_list.pop()
                            break
                    if not remove_row:
                        # increment totals for fields
                        for i, field in enumerate(display_field_paths[1:]):
                            if field in display_totals:
                                increment_total(field, display_totals, row[i + 1])
                        for position, display_property in property_list.items():
                            if not obj:
                                obj = model_class.objects.get(pk=row.pop(0))
                            relations = display_property.split('__')
                            root_relation = relations[0]
                            if root_relation in m2m_relations:
                                pk = row.pop(0)
                                if pk is not None:
                                    # a related object exists
                                    m2m_obj = getattr(obj, root_relation).get(pk=pk)
                                    val = reduce(getattr, relations[1:], m2m_obj)
                                else:
                                    val = None
                            else:
                                # Could error if a related field doesn't exist
                                try:
                                    val = reduce(getattr, relations, obj)
                                except AttributeError:
                                    val = None
                            values_and_properties_list[-1].insert(position, val)
                            increment_total(display_property, display_totals, val)
                        for position, display_custom in custom_list.items():
                            if not obj:
                                obj = model_class.objects.get(pk=row.pop(0))
                            val = obj.get_custom_value(display_custom)
                            values_and_properties_list[-1].insert(position, val)
                            increment_total(display_custom, display_totals, val)
                        filtered_report_rows += [values_and_properties_list[-1]]
                    if preview and len(filtered_report_rows) == 50:
                        break
            if hasattr(display_fields, 'filter'):
                sort_fields = display_fields.filter(
                    sort__gt=0,
                ).order_by('-sort').values_list('position', 'sort_reverse')
                for sort_field in sort_fields:
                    sort_value = sort_field[0]
                    try:
                        filtered_report_rows = sorted(
                            filtered_report_rows,
                            key=lambda x: self.sort_helper(x, sort_value),
                            reverse=sort_field[1]
                        )
                    # Crappy way to deal with null dates.
                    except TypeError:
                        try:
                            filtered_report_rows = sorted(
                                filtered_report_rows,
                                key=lambda x: self.sort_helper(
                                    x, sort_value, date_field=True),
                                reverse=sort_field[1]
                            )
                        except TypeError:
                            filtered_report_rows = sorted(
                                filtered_report_rows,
                                key=lambda x: self.sort_helper(
                                    x, sort_value, number_field=True),
                                reverse=sort_field[1]
                            )

            values_and_properties_list = filtered_report_rows
        else:
            values_and_properties_list = []
            message = "Permission Denied"

        # add choice list display and display field formatting
        choice_lists = {}
        display_formats = {}
        final_list = []
        for df in display_fields:
            if df.choices and hasattr(df, 'choices_dict'):
                df_choices = df.choices_dict
                # Insert blank and None as valid choices
                df_choices[''] = ''
                df_choices[None] = ''
                choice_lists.update({df.position: df_choices})
            if hasattr(df, 'display_format') and df.display_format:
                display_formats.update({df.position: df.display_format})

        for row in values_and_properties_list:
            # add display totals for grouped result sets
            # TODO: dry this up, duplicated logic in non-grouped total routine
            if group:
                # increment totals for fields
                for i, field in enumerate(display_field_paths[1:]):
                    if field in display_totals.keys():
                        increment_total(field, display_totals, row[i])
            row = list(row)
            for position, choice_list in choice_lists.items():
                row[position] = unicode(choice_list[row[position]])
            for position, display_format in display_formats.items():
                # convert value to be formatted into Decimal in order to apply
                # numeric formats
                try:
                    value = Decimal(row[position])
                except:
                    value = row[position]
                # Try to format the value, let it go without formatting for ValueErrors
                try:
                    row[position] = display_format.string.format(value)
                except ValueError:
                    row[position] = value
            final_list.append(row)
        values_and_properties_list = final_list

        if display_totals:
            display_totals_row = []

            fields_and_properties = list(display_field_paths[1:])
            for position, value in property_list.items():
                fields_and_properties.insert(position, value)
            for i, field in enumerate(fields_and_properties):
                if field in display_totals.keys():
                    display_totals_row += [display_totals[field]['val']]
                else:
                    display_totals_row += ['']

            # add formatting to display totals
            for df in display_fields:
                if df.display_format:
                    try:
                        value = Decimal(display_totals_row[df.position-1])
                    except:
                        value = display_totals_row[df.position-1]
                    # Fall back to original value if format string and value
                    # aren't compatible, e.g. a numerically-oriented format
                    # string with value which is not numeric.
                    try:
                        value = df.display_format.string.format(value)
                    except ValueError:
                        pass
                    display_totals_row[df.position-1] = value

            values_and_properties_list = (
                values_and_properties_list + [
                    ['TOTALS'] + (len(fields_and_properties) - 1) * ['']
                    ] + [display_totals_row]
                )

        return values_and_properties_list, message

    def sort_helper(self, x, sort_key, date_field=False, number_field=False):
        # If comparing datefields, assume null is the min year
        if date_field and x[sort_key] == None:
            result = datetime.date(datetime.MINYEAR, 1, 1)
        elif number_field and x[sort_key] == None:
            result = 0
        else:
            result = x[sort_key]
        if isinstance(result, string_types):
            return result.lower()
        elif result is None:
            return ''
        return result


class GetFieldsMixin(object):
    def get_fields(self, model_class, field_name='', path='', path_verbose=''):
        """ Get fields and meta data from a model

        :param model_class: A django model class
        :param field_name: The field name to get sub fields from
        :param path: path of our field in format
            field_name__second_field_name__ect__
        :param path_verbose: Human readable version of above
        :returns: Returns fields and meta data about such fields
            fields: Django model fields
            custom_fields: fields from django-custom-field if installed
            properties: Any properties the model has
            path: Our new path
            path_verbose: Our new human readable path
        :rtype: dict
        """
        fields = get_direct_fields_from_model(model_class)
        properties = get_properties_from_model(model_class)
        custom_fields = get_custom_fields_from_model(model_class)
        app_label = model_class._meta.app_label

        if field_name != '':
            field = model_class._meta.get_field_by_name(field_name)
            if path_verbose:
                path_verbose += "::"
            # TODO: need actual model name to generate choice list (not pluralized field name)
            # - maybe store this as a separate value?
            if field[3] and hasattr(field[0], 'm2m_reverse_field_name'):
                path_verbose += field[0].m2m_reverse_field_name()
            else:
                path_verbose += field[0].name

            path += field_name
            path += '__'
            if field[2]:  # Direct field
                try:
                    new_model = field[0].related.parent_model
                except AttributeError:
                    new_model = field[0].related.model
                path_verbose = new_model.__name__.lower()
            else:  # Indirect related field
                try:
                    new_model = field[0].related_model
                except AttributeError:  # Django 1.7
                    new_model = field[0].model
                path_verbose = new_model.__name__.lower()

            fields = get_direct_fields_from_model(new_model)

            custom_fields = get_custom_fields_from_model(new_model)
            properties = get_properties_from_model(new_model)
            app_label = new_model._meta.app_label

        return {
            'fields': fields,
            'custom_fields': custom_fields,
            'properties': properties,
            'path': path,
            'path_verbose': path_verbose,
            'app_label': app_label,
        }

    def get_related_fields(self, model_class, field_name, path="", path_verbose=""):
        """ Get fields for a given model """
        if field_name:
            field = model_class._meta.get_field_by_name(field_name)
            if field[2]:
                # Direct field
                try:
                    new_model = field[0].related.parent_model()
                except AttributeError:
                    new_model = field[0].related.model
            else:
                # Indirect related field
                new_model = field[0].model()

            if path_verbose:
                path_verbose += "::"
            path_verbose += field[0].name

            path += field_name
            path += '__'
        else:
            new_model = model_class

        new_fields = get_relation_fields_from_model(new_model)
        model_ct = ContentType.objects.get_for_model(new_model)

        return (new_fields, model_ct, path)
