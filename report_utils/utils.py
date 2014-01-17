try:
    import cStringIO as StringIO
except ImportError:
    import StringIO
    
from django.http import HttpResponse
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.cell import get_column_letter
import re

class DataExportMixin(object):
    def list_to_xlsx_response(self, data, title='report', header=None):
        """ Make 2D list into a xlsx response for download """
        wb = Workbook()
        ws = wb.worksheets[0]
        title = re.sub(r'\W+', '', title)[:30]
        ws.title = title
        if not title.endswith('.xlsx'):
            title += '.xlsx'
            
        if header:
            i = 0
            for header_cell in header:
                cell = ws.cell(row=0, column=i)
                cell.value = header_cell
                cell.style.font.bold = True
                ws.column_dimensions[get_column_letter(i+1)].width = field.width
                i += 1
        
        for row in data:
            try:
                ws.append(row)
            except ValueError as e:
                ws.append([e.message])
            except:
                ws.append(['Unknown Error'])
        
        myfile = StringIO.StringIO()
        myfile.write(save_virtual_workbook(wb))
        response = HttpResponse(
            myfile.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s' % title
        response['Content-Length'] = myfile.tell()
        return response

